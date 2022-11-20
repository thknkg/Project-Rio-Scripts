import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def export_to_excel(stats, excel_file):
    wb = load_workbook(excel_file)
    sheets = wb.sheetnames
    ws1 = wb[sheets[0]]

    try:
        del ws1.tables["MarioBaseballStats"]
    except KeyError:
        table = True

    wb.save(excel_file)

    ws1.delete_rows(1, ws1.max_row)  # clears out old values in the excel sheet
    wb.save(excel_file)

    for row in stats:
        ws1.append(row)
    wb.save(excel_file)

    def column_string(n):
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string

    table_size = "A1:" + str(column_string(len(stats[0]))) + str(len(stats))
    tab = Table(displayName="MarioBaseballStats", ref=table_size)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws1.add_table(tab)
    wb.save(excel_file)